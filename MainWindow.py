import requests
from PyQt5.QtWidgets import QMainWindow, QFileDialog
from PyQt5 import uic
from AuthWindow import AuthWindow, Cache
import json
import os
from datetime import datetime, date, timedelta
from openpyxl import load_workbook
import pandas as pd
from dateutil.relativedelta import relativedelta
import csv
import sys
import xlwings as xw
from decimal import Decimal
import psycopg2
from psycopg2 import Error


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        main_path = self.resource_path('static/main.ui')
        uic.loadUi(main_path, self)
        current_date = str(date.today())
        start = current_date + ' ' + '00:00:00'
        end = current_date + ' ' + '23:59:59'
        start = datetime.strptime(start, '%Y-%m-%d %H:%M:%S')
        end = datetime.strptime(end, '%Y-%m-%d %H:%M:%S')
        self.dateTimeEdit.setDateTime(start)
        self.dateTimeEdit_2.setDateTime(end)
        self.action_2.triggered.connect(self.create_window)
        self.pushButton.clicked.connect(self.upload_excel)
        self.pushButton_2.clicked.connect(self.upload_excel2)
        self.pushButton_3.clicked.connect(self.upload_excel3)
        self.pushButton_4.clicked.connect(self.upload_excel4)
        self.pushButton_5.clicked.connect(self.upload_excel5)
        self.pushButton_6.clicked.connect(self.upload_excel6)
        self.pushButton_7.clicked.connect(self.upload_excel7)

        self.start_1.clicked.connect(self.start_app1)
        self.start_2.clicked.connect(self.start_app2)
        self.start_3.clicked.connect(self.start_app3)
        self.start_4.clicked.connect(self.start_app4)
        self.start_5.clicked.connect(self.start_app5)
        self.start_6.clicked.connect(self.start_app6)
        self.start_7.clicked.connect(self.start_app7)
        self.start_8.clicked.connect(self.start_app8)
        self.start_9.clicked.connect(self.start_app9)
        self.start_10.clicked.connect(self.start_app10)
        self.start_11.clicked.connect(self.start_app11)
        self.start_14.clicked.connect(self.start_app14)
        self.start_15.clicked.connect(self.start_app15)
        # Надо использовать QtWidget.setToolTip('text')

    def create_window(self):
        window = AuthWindow(self)
        window.show()

    @staticmethod
    def resource_path(relative_path):
        try:
            base_path = sys._MEIPASS
        except Exception:
            base_path = os.path.abspath(".")
        return os.path.join(base_path, relative_path)

    def save_log(self, text):
        self.logs.setReadOnly(False)
        self.logs.appendPlainText(text)
        self.logs.setReadOnly(True)

    def path_file(self, label, label2=None, label3=None, filetype=0):
        self.logs.clear()
        self.save_log('Идёт чтение файла')
        filetypes = [("Excel (*.xlsx *.xls *.xlsm)"), ("txt (*.txt)"), ("csv (*.csv)")]
        path = QFileDialog.getOpenFileName(self, 'Выбрать файл', '',filetypes[filetype])[0]
        file_name = os.path.basename(path)
        if path == "":
            self.logs.clear()
            self.save_log(text='Вы не выбрали файл')
            label.setText('')
            if label2 is None:
                pass
            else:
                label2.setText('')
                label3.setText('')
        else:
            label.setText(path)
            self.logs.clear()
            self.save_log(text='Вы выбрали файл: ' + file_name)
        return path

    def open_excel(self, label, label2, label3):
        path = self.path_file(label, label2, label3)
        if path == "":
            return
        else:
            try:
                wb = load_workbook(path)
                sheets = wb.sheetnames
                sheet_row = []
                for sheet in sheets:
                    df = pd.read_excel(io=path, sheet_name=sheet, header=None)
                    count_columns = len(df.axes[1])
                    array = []
                    for i in range(count_columns):
                        count_rows = df[df.columns[i]].count()
                        array.append({i: count_rows})
                    sheet_row.append({sheet: array})
                label2.setText(str(sheets))
                label3.setText(str(sheet_row))
            except Exception:
                self.logs.clear()
                label.setText('')
                label2.setText('')
                label3.setText('')
                self.save_log('Выберите корректный файл')

    def upload_excel(self):
        self.open_excel(label=self.label_9, label2=self.label_10, label3=self.label_11)

    def upload_excel2(self):
        self.open_excel(label=self.label_20, label2=self.label_21, label3=self.label_22)

    def upload_excel3(self):
        self.open_excel(label=self.label_31, label2=self.label_32, label3=self.label_33)

    def upload_excel4(self):
        self.open_excel(label=self.label_40, label2=self.label_41, label3=self.label_42)

    def upload_excel5(self):
        self.path_file(self.label_44, filetype=1)

    def upload_excel6(self):
        self.path_file(self.label_46, filetype=1)

    def upload_excel7(self):
        self.path_file(self.label_48)

    def validate_integer(self, value):
        try:
            value = int(value)
            return value
        except Exception:
            return 'Invalid'

    @staticmethod
    def show_input(label):
        result = label.text()
        return result

    def sheets_excel(self, label):
        path = self.show_input(label)
        wb = load_workbook(path)
        sheets = wb.sheetnames
        return sheets

    def validate_input_sheet(self, label, line):
        sheets = self.sheets_excel(label)
        input_sheet_null = self.show_input(line)
        input_sheet = self.validate_integer(self.show_input(line))
        if input_sheet_null == '':
            self.save_log('Вы не выбрали лист')
            return False
        elif input_sheet is False:
            self.save_log('Вы ввели некорректное значение')
            return False
        elif input_sheet >= len(sheets):
            self.save_log('Вы выбрали не существующий лист')
            return False
        else:
            return sheets[self.validate_integer(input_sheet)]

    def validate_input_columns(self, label, line, line2):
        columns = self.show_input(line2)
        if columns == '':
            return columns
        else:
            columns = [self.validate_integer(column) for column in columns.split(',')]
            if 'Invalid' in columns:
                self.save_log('Вы ввели некорректное значение в поле "Выбрать столбцы"')
                return False
            else:
                path = self.show_input(label)
                sheets = self.sheets_excel(label)
                sheet = sheets[int(self.show_input(line))]
                df = pd.read_excel(path, sheet_name=sheet, header=None)
                all_columns = [i for i in range(len(df.axes[1]))]
                bool_columns = tuple(x in all_columns for x in columns)
                if False in bool_columns:
                    self.save_log('Вы выбрали несуществующий столбец в поле "Выбрать столбцы"')
                    return False
                else:
                    return columns

    def validate_input_slice(self, line):
        if self.validate_integer(self.show_input(label=line)) == 'Invalid' or self.validate_integer(
                self.show_input(label=line)) == 0:
            self.save_log('Вы не выбрали по сколько разделить или ввели некорректное значение')
            return False
        else:
            return self.validate_integer(self.show_input(label=line))

    def validate_input(self, label, line, line2, line3):
        path = self.show_input(label=label)
        if path == '':
            self.save_log('Вы не выбрали файл')
        else:
            input_sheet = self.validate_input_sheet(label, line)
            input_columns = self.validate_input_columns(label, line, line2)
            input_slice = self.validate_input_slice(line3)
            if input_sheet is False:
                pass
            elif input_columns is False:
                pass
            elif input_slice is False:
                pass
            else:
                self.save_log('Вы выбрали лист: ' + str(input_sheet))
                if input_columns == '':
                    self.save_log('Вы выбрали все столбцы')
                else:
                    self.save_log('Вы выбрали столбцы: ' + str(', '.join(map(str, input_columns))))
                self.save_log('Вы выбрали разделить по: ' + str(input_slice))
                return True

    def vpn_on(self):
        try:
            requests.get('https://order-backoffice-apigateway.samokat.ru/swagger-ui/index.html?'
                         'configUrl=%2Fv3%2Fapi-docs%2Fswagger-config&urls.primaryName=backoffice-public-api')
            return True
        except Exception:
            self.save_log('Включите vpn')
            return False

    def get_dataframe(self, label, line, line2):
        sheets = self.sheets_excel(label)
        sheet = sheets[int(self.show_input(line))]
        path = self.show_input(label)
        dfs = pd.read_excel(path, sheet_name=sheet, header=None, dtype=str)
        df = pd.DataFrame([])
        columns = self.show_input(line2)
        if columns == '':
            for i in range(len(dfs.axes[1])):
                df = pd.concat([df, dfs[i]], ignore_index=True)
            df = df.dropna(axis=0, how='any')
            df = df.reset_index(drop=True)
            return df
        else:
            columns = [int(column) for column in columns.split(',')]
            for i in range(len(dfs.axes[1])):
                for column in columns:
                    if i == column:
                        df = pd.concat([df, dfs[i]], ignore_index=True)
            df = df.dropna(axis=0, how='any')
            df = df.reset_index(drop=True)
            return df

    def validate_df(self, df):
        symbols = [',', ';', ':', ' ', '\\.', '\\(', '\\)']
        for symbol in symbols:
            df = df.replace(symbol, '', regex=True)
        self.save_log('Файл провалидирован')
        return df

    def find_duplicates(self, df):
        duplicates = df[df.duplicated()]
        duplicates = duplicates.drop_duplicates()
        duplicates = duplicates[0].values.astype(str).tolist()
        return duplicates

    def drop_duplicates(self, df):
        df = df.drop_duplicates()
        self.save_log('Дубликаты удалены')
        return df

    def df_slice(self, df, line, vtype):
        result = []
        _max = len(df.axes[0])
        n = int(self.show_input(line))
        for start in range(0, _max, n):
            stop = start + n
            slice_object = slice(start, stop)
            result.append(df[slice_object][0].values.astype(vtype).tolist())
        return result

    def create_csv(self, checkbox, result, vtype):
        if checkbox.isChecked() is True:
            str_current_datetime = str(datetime.now()).replace(':', '-')
            if vtype == str:
                _name = "promocodes"
                delimiter = '","'
            elif vtype == int:
                _name = "users"
                delimiter = ','
            file_name = "postman(" + _name + ")" + str_current_datetime + ".csv"
            with open(file_name, "w", encoding='utf-8', newline="") as f:
                writer = csv.writer(f)
                writer.writerow([_name])
                for i in result:
                    i = delimiter.join([str(n) for n in i])
                    writer.writerow([i])
                self.save_log('Готово, создан файл: ' + file_name)

    def start_app1(self):
        try:
            valid = self.validate_input(label=self.label_9, line=self.lineEdit, line2=self.lineEdit_2,
                                        line3=self.lineEdit_3)
            if valid is True:
                df = self.validate_df(self.get_dataframe(label=self.label_9, line=self.lineEdit, line2=self.lineEdit_2))
                duplicates = self.find_duplicates(df)
                df = self.drop_duplicates(df)
                result = self.df_slice(df, self.lineEdit_3, str)
                str_current_datetime = str(datetime.now()).replace(':', '-')
                file_name = "promotion(promocodes) " + str_current_datetime + ".json"
                with open(file_name, 'w', encoding='utf-8') as file:
                    for i in result:
                        file.write(
                            f'{{"promotionId": "{self.lineEdit_4.text()}", "promocodes": {i}}}\n')
                    file.write(f'Дубликаты {duplicates}\n')
                with open(file_name, 'r', encoding='utf-8') as f:
                    old_data = f.read()
                new_data = old_data.replace("'", '"')
                new_data = new_data.replace('\\xa0', '')
                with open(file_name, 'w', encoding='utf-8') as f:
                    f.write(new_data)
                    self.save_log('Готово, создан файл: ' + file_name)
                self.create_csv(self.checkBox, result, str)
        except Exception:
            self.save_log('Что-то пошло не так')

    def start_app2(self):
        try:
            valid = self.validate_input(label=self.label_20, line=self.lineEdit_5, line2=self.lineEdit_6,
                                        line3=self.lineEdit_7)
            if valid is True:
                df = self.validate_df(self.get_dataframe(label=self.label_20, line=self.lineEdit_5, line2=self.lineEdit_6))
                duplicates = self.find_duplicates(df)
                df = self.drop_duplicates(df)
                result = self.df_slice(df, self.lineEdit_7, int)
                str_current_datetime = str(datetime.now()).replace(':', '-')
                file_name = "promotion(users) " + str_current_datetime + ".json"
                with open(file_name, 'w', encoding='utf-8') as file:
                    for i in result:
                        file.write(
                            f'{{"promotionId": "{self.show_input(self.lineEdit_8)}", "userIds": {i}, '
                            f'"userType": "SAMOKAT", "disableNotifications": true}}\n')
                    file.write(f'Дубликаты {duplicates}\n')
                with open(file_name, 'r', encoding='utf-8') as f:
                    old_data = f.read()
                new_data = old_data.replace('\\xa0', '')
                with open(file_name, 'w', encoding='utf-8') as f:
                    f.write(new_data)
                    self.save_log('Готово, создан файл: ' + file_name)
                self.create_csv(self.checkBox_2, result, int)
        except Exception:
            self.save_log('Что-то пошло не так')

    def start_app3(self):
        try:
            valid = self.validate_input(label=self.label_31, line=self.lineEdit_9, line2=self.lineEdit_10,
                                        line3=self.lineEdit_11)
            if valid is True:
                df = self.validate_df(self.get_dataframe(label=self.label_31, line=self.lineEdit_9,
                                                         line2=self.lineEdit_10))
                duplicates = self.find_duplicates(df)
                df = self.drop_duplicates(df)
                result = self.df_slice(df, self.lineEdit_11, int)
                str_current_datetime = str(datetime.now()).replace(':', '-')
                file_name = "banner(users) " + str_current_datetime + ".json"
                with open(file_name, 'w', encoding='utf-8') as file:
                    for i in result:
                        file.write(
                            f'{{"userIds": {i}, "userType": "SAMOKAT", "bannerId": '
                            f'"{self.show_input(self.lineEdit_12)}"}}\n')
                    file.write(f'Дубликаты {duplicates}\n')
                with open(file_name, 'r', encoding='utf-8') as f:
                    old_data = f.read()
                new_data = old_data.replace('\\xa0', '')
                with open(file_name, 'w', encoding='utf-8') as f:
                    f.write(new_data)
                    self.save_log('Готово, создан файл: ' + file_name)
                self.create_csv(self.checkBox_3, result, int)
        except Exception:
            self.save_log('Что-то пошло не так')

    def start_app4(self):
        try:
            valid = self.validate_input(label=self.label_40, line=self.lineEdit_13, line2=self.lineEdit_14,
                                        line3=self.lineEdit_15)
            if valid is True:
                df = self.validate_df(self.get_dataframe(label=self.label_40, line=self.lineEdit_13,
                                                         line2=self.lineEdit_14))
                duplicates = self.find_duplicates(df)
                df = self.drop_duplicates(df)
                result = self.df_slice(df, self.lineEdit_15, int)
                str_current_datetime = str(datetime.now()).replace(':', '-')
                file_name = "newfile " + str_current_datetime + ".json"
                with open(file_name, 'w', encoding='utf-8') as file:
                    for i in result:
                        file.write(f'{i}\n')
                    file.write(f'Дубликаты {duplicates}\n')
                with open(file_name, 'r', encoding='utf-8') as f:
                    old_data = f.read()
                new_data = old_data.replace("'", '')
                new_data = new_data.replace('\\xa0', '')
                with open(file_name, 'w', encoding='utf-8') as f:
                    f.write(new_data)
                    self.save_log('Готово, создан файл: ' + file_name)
        except Exception:
            self.save_log('Что-то пошло не так')


    def start_app5(self):
        path = self.show_input(self.label_44)
        vpn = self.vpn_on()
        if vpn:
            if path == '':
                self.save_log('Вы не выбрали файл')
            else:
                try:
                    text = open(path, "r", encoding="utf-8")
                    text = text.read().split('\n')
                    result, results = [], []
                    for js in text:
                        jsons = json.loads(js)
                        for shipment in jsons['SHIPMENTS']:
                            items, temp, temp_arr = [], [], {}
                            names = [item['ITEM'] for item in shipment['DETAIL']]
                            res_slice, responses = [], {'data': []}
                            for start in range(0, len(names), 2):
                                stop = start + 200
                                slice_object = slice(start, stop)
                                res_slice.append(names[slice_object])
                            for i in res_slice:
                                search = {"filter": {"nomenclatureCodes": i}, "limit": len(i)}
                                response = requests.post('https://ds-metadata.samokat.ru/products/by-filter',
                                                         json=search)
                                response_json = response.json()
                                responses['data'].extend(response_json['data'])
                            for item, name in zip(shipment['DETAIL'], names):
                                if names.count(name) > 1:
                                    if name in [key for key, value in temp_arr.items()]:
                                        temp_arr[name].append(item)
                                    else:
                                        temp_arr.update({name: [item]})
                                else:
                                    valid_packages = []
                                    packages = [i['packages'] for i in responses['data'] if i['nomenclatureCode'] == name][0]
                                    for package in packages:
                                        if package['packageType'].upper() == item['OP_QTY_UM'].upper():
                                            quantity = item['QUANTITY']
                                            op_qty = item['OP_QTY']
                                            if isinstance(item['QUANTITY'], float) == False:
                                                quantity = Decimal(item['QUANTITY'].split('.')[0]).quantize(Decimal("1.00"))
                                                op_qty = Decimal(item['OP_QTY'].split('.')[0]).quantize(Decimal("1.00"))

                                            coefficient = package['coefficient']
                                            if coefficient * op_qty == quantity:
                                                valid_packages.append(True)
                                            else:
                                                valid_packages.append(False)

                                    if True not in valid_packages:
                                        pack = {
                                            'packages': [package['name'].replace('\xa0', '') for package in packages]}
                                        item.update(pack)
                                        items.append(item)

                            for e in temp_arr:
                                quantity = 0
                                op_qty = 0
                                valid_packages = []
                                packages = [i['packages'] for i in responses['data'] if i['nomenclatureCode'] == e][0]
                                for i in temp_arr[e]:
                                    quantity += Decimal(i['QUANTITY']).quantize(Decimal("1.00"))
                                    op_qty += Decimal(i['OP_QTY']).quantize(Decimal("1.00"))
                                    for package in packages:
                                        if package['packageType'] == i['OP_QTY_UM']:
                                            coefficient = package['coefficient']
                                            if coefficient * op_qty == quantity:
                                                valid_packages.append(True)
                                            else:
                                                valid_packages.append(False)
                                if True not in valid_packages:
                                    pack = {
                                        'packages': [package['name'].replace('\xa0', '') for package in
                                                     packages]}
                                    temp_arr[e].append(pack)
                                    items.append(temp_arr[e])
                            for x in items:
                                if x not in temp:
                                    temp.append(x)
                            items = temp
                            result.append({shipment['SHIPMENT_ID']: items})

                    for res in result:
                        for key, value in res.items():
                            if value == []:
                                continue
                            else:
                                res = res[list(res.keys())[0]]
                                res.insert(0,{'count': len(res)})
                                results.append(res)
                    str_current_datetime = str(datetime.now()).replace(':', '-')
                    file_name = "shipment(YT) " + str_current_datetime + ".sql"
                    with open(file_name, 'w', encoding="utf-8") as f:
                        f.write(json.dumps(results, indent=4, ensure_ascii=False))
                        self.save_log('Готово, создан файл: ' + file_name)
                except Exception:
                    self.save_log('Некорректный JSON или что-то пошло не так')
    def start_app6(self):
        path = self.show_input(self.label_46)
        if path == '':
            self.save_log('Вы не выбрали файл')
        else:
            try:
                path = self.show_input(self.label_46)
                jsons = open(path, "r", encoding="utf-8")
                jsons = json.loads(jsons.read())
                result = []
                data = date.today()
                if 'RECEIPTS' in jsons:
                    confirm = 'RECEIPTS'
                elif 'SHIPMENTS' in jsons:
                    confirm = 'SHIPMENTS'
                length2 = len(jsons[confirm][0]['DETAIL'])
                for n in range(0, length2 - 1):
                    if jsons[confirm][0]['DETAIL'][n]['MAN_DATE'] == '0001-01-01':
                        result.append(jsons[confirm][0]['DETAIL'][n])
                    elif jsons[confirm][0]['DETAIL'][n]['MAN_DATE'] > str(data):
                        result.append(jsons[confirm][0]['DETAIL'][n])
                    elif jsons[confirm][0]['DETAIL'][n]['EXP_DATE'] < str(data):
                        result.append(jsons[confirm][0]['DETAIL'][n])
                    elif jsons[confirm][0]['DETAIL'][n]['MAN_DATE'] < str(data - relativedelta(years=10)):
                        result.append(jsons[confirm][0]['DETAIL'][n])
                    elif jsons[confirm][0]['DETAIL'][n]['EXP_DATE'] > str(data + relativedelta(years=15)):
                        result.append(jsons[confirm][0]['DETAIL'][n])
                str_current_datetime = str(datetime.now()).replace(':', '-')
                file_name = "shipment(date) " + str_current_datetime + ".sql"
                with open(file_name, 'w', encoding='utf-8') as file:
                    file.write(f'{result}\n')
                with open(file_name, 'r', encoding='utf-8') as f:
                    old_data = f.read()
                new_data = old_data.replace(',', ', \n')
                with open(file_name, 'w', encoding='utf-8') as f:
                    f.write(new_data)
                    self.save_log('Готово, создан файл: ' + file_name)
            except Exception:
                self.save_log('Некорректный JSON')


    def start_app7(self):
        self.logs.clear()
        path = self.show_input(self.label_48)
        if path == '':
            self.save_log('Вы не выбрали файл')
        else:
            vpn = self.vpn_on()
            if vpn is True:
                try:
                    df = pd.read_excel(path)
                    str_current_datetime = str(datetime.now()).replace(':', '-')
                    file_name = 'manual_shipments ' + str_current_datetime + '.json'
                    shipment_json = [{"shipmentId": "Вставить id перемещения",
                                      "documentNumber": "Вставить номер перемещения", "products": []}]
                    products = [product for product in df['productId']]
                    search_json = {"productIds": products}
                    response = requests.post('https://ds-metadata.samokat.ru/products/by-ids', json=search_json)
                    response_json = response.json()
                    count_YT = 0
                    for product, quantity, date_1, date_2 in zip(df['productId'], df['totalProductQuantity'],
                                                                 df['productionDate'], df['bestBeforeDate']):
                        date_1 = str(date_1).split(' ')[0] + "T00:00:00.00Z"
                        date_2 = str(date_2).split(' ')[0] + "T00:00:00.00Z"
                        try:
                            product_yt = [p['nomenclatureCode'] for p in response_json if p['productId'] == product][0]
                        except Exception:
                            count_YT += 1
                            product_yt = 'Not Found'
                        product_json = {
                            "productId": product,
                            "productCode": product_yt,
                            "totalProductQuantity": quantity,
                            "packages": [
                                {
                                    "productQuantity": quantity,
                                    "productsPerPackageCoefficient": 1
                                }
                            ],
                            "productionDate": date_1,
                            "bestBeforeDate": date_2,
                            "isDamaged": False
                        }
                        shipment_json[0]['products'].append(product_json)
                    with open(file_name, 'w', encoding="utf-8") as f:
                        f.write(f'Кол-во не найденных УТ: {count_YT}\n')
                        f.write(json.dumps(shipment_json, indent=4, ensure_ascii=False))
                    self.save_log('Готово, создан файл: ' + file_name)
                except Exception:
                    self.save_log('Некорректный excel файл')

    def start_app8(self):
        self.logs.clear()
        vpn = self.vpn_on()
        if vpn is True:
            try:
                token = Cache.load("token")
                guids = self.plainTextEdit.toPlainText().split('\n')
                guids = [guid for guid in guids if len(guid) == 36]
                if len(guids) == 0:
                    self.save_log('Вы не ввели guid ЦФЗ')
                else:
                    self.save_log('Вы ввели ' + str(len(guids)) + ' guid ЦФЗ')
                    invalid_guids = []
                    result,datas, showcases = [], {}, []
                    count_showcases = 0
                    count_receipts = 0
                    for guid in guids:
                        if len(guid) == 36:
                            cfz_setting = []
                            showcases_search = {"storeId": guid}
                            receipts_search = {"providerId": 0, "warehouseId": guid}
                            header = {'Authorization': 'Bearer ' + token}
                            url_showcases = 'https://order-backoffice-apigateway.samokat.ru/showcases/getBy'
                            url_receipts = 'https://smk-supportpaymentgw.samokat.ru/receipt/cash-registers/find'
                            showcase = requests.post(url_showcases, headers=header, json=showcases_search)
                            receipt = requests.get(url_receipts, headers=header, params=receipts_search)
                            showcase = showcase.json()

                            if showcase == {"items": []}:
                                pass
                            else:
                                count_showcases += 1
                                showcases.append(showcase['items'][0]['showcaseId'])
                            receipt = receipt.json()
                            if receipt == {"error": "NOT_FOUND", "value": None}:
                                receipts_search = {"providerId": 2, "warehouseId": guid}
                                receipt = requests.get(url_receipts, headers=header, params=receipts_search)
                                receipt = receipt.json()
                                if receipt == {"error": "NOT_FOUND", "value": None}:
                                    pass
                                else:
                                    count_receipts += 1
                            else:
                                count_receipts += 1
                            cfz_setting.append(showcase)
                            cfz_setting.append(receipt)
                            result.append({guid: cfz_setting})

                            if self.checkBox_4.isChecked() is False:
                                pass
                            else:
                                cfz = requests.get(f'https://ds-warehouse.samokat.ru/warehouses/load/by-id/{guid}')
                                cfz = cfz.json()
                                cfz = cfz['value']
                                cityId = cfz['cityId']
                                city = requests.get(f'https://ds-warehouse.samokat.ru/city/{cityId}')
                                city = city.json()
                                city = city['value']
                                show, group_code, login, password = '', '', '', ''
                                if showcase == {"items": []}:
                                    pass
                                else:
                                    show = 'Экспресс'
                                if receipt == {"error": "NOT_FOUND", "value": None}:
                                    receipts_search = {"providerId": 2, "warehouseId": guid}
                                    receipt = requests.get(url_receipts, headers=header, params=receipts_search)
                                    receipt = receipt.json()
                                    if receipt == {"error": "NOT_FOUND", "value": None}:
                                        pass
                                    else:
                                        group_code = receipt['value']['cashRegisterGroup']
                                        login = receipt['value']['login']
                                        password = receipt['value']['password']
                                else:
                                    group_code = receipt['value']['cashRegisterGroup']
                                    login = receipt['value']['login']
                                    password = receipt['value']['password']

                                data = [cfz['name'], cfz['startedToOperate'].split('T')[0], cfz['warehouseId'],
                                        cfz['email'], city['code'], cfz['address']['lat'], cfz['address']['lon'],
                                        cfz['address']['fullAddress'], show, group_code, login, password]
                                datas[guid] = data
                        else:
                            invalid_guids.append(guid)
                    if invalid_guids == [] or invalid_guids == ['']:
                        pass
                    else:
                        self.save_log(
                            'В списке присутсвуют некорректные guid: ' + str(', '.join(map(str, invalid_guids))))
                    str_current_datetime = str(datetime.now()).replace(':', '-')
                    file_name = 'cfz_settings ' + str_current_datetime + '.json'
                    with open(file_name, 'w', encoding="utf-8") as f:
                        f.write(f'Кол-во витрин: {count_showcases}\n')
                        f.write(f'Кол-во касс: {count_receipts}\n')
                        f.write(f'Список витрин:\n')
                        for show in showcases:
                            f.write(f'{show}\n')
                        f.write(f'Список фидов:\n')
                        for show in showcases:
                            f.write(f'https://partners-api.samokat.ru/showcase/v2/search/support/feed/{show}\n')
                        f.write(json.dumps(result, indent=4, ensure_ascii=False))
                    self.save_log('Готово, создан файл: ' + file_name)
                    if self.checkBox_4.isChecked() is True:
                        df = pd.DataFrame(datas)
                        file_name2 = 'cfz_settings ' + str_current_datetime + '.xlsx'
                        writer = pd.ExcelWriter(file_name2)
                        df.to_excel(writer, index=False, header=False)
                        writer.close()
                        self.save_log('Готово, создан файл: ' + file_name2)

            except Exception:
                self.save_log('Вы не авторизовались')

    def start_app9(self):
        self.logs.clear()
        vpn = self.vpn_on()
        if vpn is True:
            try:
                order_id = self.plainTextEdit_2.toPlainText().split('\n')
                if order_id == ['']:
                    self.save_log('Вы не ввели order id')
                else:
                    if '' in order_id:
                        order_id.remove('')
                    order_id = (','.join(["'" + i + "'" for i in order_id]))
                    login = Cache.load("login")
                    password = Cache.load("password")
                    if login is None:
                        self.save_log('Вы не авторизовались')
                    else:
                        connection = psycopg2.connect(user=login,
                                                      password=password,
                                                      host="patroni-17.samokat.io",
                                                      port="5434",
                                                      dbname="order_history")
                        cursor = connection.cursor()
                        cursor.execute(f'WITH orders AS (SELECT order_id, order_line_changed  FROM order_history WHERE order_id in ({order_id})), date_created AS (SELECT change_date as created_date, order_id FROM order_status WHERE order_id in (SELECT order_id FROM orders) and order_status_id = 3), date_picking AS (SELECT change_date as picking_date, order_id FROM order_status WHERE order_id in (SELECT order_id FROM orders) and order_status_id = 4), date_picking_hub AS (SELECT change_date as picking_date_hub, order_id FROM order_status WHERE order_id in (SELECT order_id FROM orders) and order_status_id = 9), date_picked AS (SELECT change_date as picked_date,order_id FROM order_status WHERE order_id in (SELECT order_id FROM orders) and order_status_id = 5), date_picked_hub AS (SELECT change_date as picked_date_hub,order_id FROM order_status WHERE order_id in (SELECT order_id FROM orders) and order_status_id = 10), date_change AS (SELECT created_date_time as change_date, order_id FROM order_change WHERE order_id in (SELECT order_id FROM orders)), hub_picker AS (SELECT order_id, picker_id FROM distribution_center_picking_info WHERE order_id in (SELECT order_id FROM orders)), cfz_picker AS (SELECT order_id, picker_uuid FROM picking_info WHERE order_id in (SELECT order_id FROM orders)), cfz_deliveryman AS (SELECT order_id, deliveryman_uuid FROM delivery_info WHERE order_id in (SELECT order_id FROM orders)), order_number AS (SELECT order_id, display_number FROM order_history WHERE order_id in (SELECT order_id FROM orders)), product AS (SELECT oc.created_date_time as change_dates, olc.product_id FROM order_change oc JOIN order_line_change olc ON olc.order_change_id = oc.id WHERE oc.order_id in (SELECT order_id FROM orders)), accepted AS(SELECT order_id, product_id, quantity FROM order_line ol JOIN accepted_order_line aol ON aol.order_line_id = ol.id WHERE aol.order_id in (SELECT order_id FROM orders)), actual AS(SELECT order_id, product_id, quantity FROM order_line ol JOIN actual_order_line aol ON aol.order_line_id = ol.id WHERE aol.order_id in (SELECT order_id FROM orders)), changes AS(SELECT accepted.order_id, accepted.product_id, CASE WHEN accepted.quantity > actual.quantity THEN true ELSE false END AS change FROM accepted JOIN actual ON actual.product_id = accepted.product_id and actual.order_id = accepted.order_id), changes_2 AS (SELECT order_id, product_id FROM changes WHERE change is true), result AS(SELECT order_number.display_number AS "Номер заказа", to_char(date_created.created_date, \'yyyy-mm-dd hh24:mi\') AS "Время заказа", date_created.order_id, CASE WHEN change_date < picking_date_hub THEN true WHEN change_date < picking_date and picking_date_hub is NULL THEN true WHEN change_date is NULL and order_line_changed = true THEN true WHEN change_date is not NULL and order_line_changed = true and picking_date_hub is NULL and picking_date is NULL THEN true ELSE false END "Автокорректировка", CASE WHEN change_date > picking_date_hub and change_date < picked_date_hub THEN true ELSE false END "Сборка на Хабе", hub_picker.picker_id, CASE WHEN change_date > picking_date and change_date < picked_date and picking_date_hub is NULL THEN true ELSE false END "Сборка на ЦФЗ", cfz_picker.picker_uuid, CASE WHEN change_date > picked_date THEN true ELSE false END "Доставка", cfz_deliveryman.deliveryman_uuid, product.product_id AS "Продукт", changes_2.product_id AS "Продукт2" FROM date_created LEFT JOIN date_picking ON date_created.order_id = date_picking.order_id LEFT JOIN date_picking_hub ON date_picking_hub.order_id = date_created.order_id LEFT JOIN date_picked ON date_created.order_id = date_picked.order_id LEFT JOIN date_picked_hub ON date_picked_hub.order_id = date_created.order_id LEFT JOIN date_change ON date_change.order_id = date_created.order_id LEFT JOIN orders ON orders.order_id = date_created.order_id LEFT JOIN hub_picker ON hub_picker.order_id = date_created.order_id LEFT JOIN cfz_picker ON cfz_picker.order_id = date_created.order_id LEFT JOIN cfz_deliveryman ON cfz_deliveryman.order_id = date_created.order_id LEFT JOIN product ON product.change_dates = date_change.change_date LEFT JOIN order_number ON order_number.order_id = date_created.order_id LEFT JOIN changes_2 ON changes_2.order_id = date_created.order_id GROUP BY "Время заказа", date_created.order_id, "Автокорректировка", "Сборка на Хабе","Сборка на ЦФЗ", "Доставка", hub_picker.picker_id, cfz_picker.picker_uuid,cfz_deliveryman.deliveryman_uuid, "Продукт", "Номер заказа", "Продукт2" ORDER BY date_created.order_id ASC, "Сборка на Хабе" DESC) SELECT "Номер заказа", "Время заказа", order_id, "Автокорректировка", "Сборка на Хабе", picker_id, "Сборка на ЦФЗ", picker_uuid, "Доставка", deliveryman_uuid, CASE  WHEN "Автокорректировка" is true and "Продукт2" is NULL THEN "Продукт" WHEN "Автокорректировка" is true and "Продукт2" is not NULL THEN "Продукт2" WHEN "Автокорректировка" is false THEN "Продукт" END "Продукт" FROM result')
                        result = []
                        result.extend(cursor.fetchall())
                        if not result:
                            self.save_log('Корректировки отсутствуют')
                        else:
                            df = pd.DataFrame(result)
                            type_update = []
                            for a,b,c,d in zip(df[3].tolist(),df[4].tolist(),df[6].tolist(),df[8].tolist()):
                                if a is True:
                                    type_update.append('Автокорректировка')
                                if b is True:
                                    type_update.append('Ручная(Сборка на ХАБе)')
                                if c is True:
                                    type_update.append('Ручная(Сборка на ЦФЗ)')
                                if d is True:
                                    type_update.append('Ручная(На этапе доставки)')

                            result = {'Номер заказа': df[0].tolist(), 'Время заказа': df[1].tolist(), 'order_id': df[2].tolist(), 'Тип корректировки': type_update, 'product_id': df[10].tolist()}
                            res = pd.DataFrame(result)
                            who = []
                            for type, uuid1, uuid2, uuid3 in zip(res['Тип корректировки'].tolist(), df[5].tolist(), df[7].tolist(), df[9].tolist()):
                                if type == 'Автокорректировка':
                                    who.append('')
                                if type == 'Ручная(Сборка на ХАБе)':
                                    who.append(uuid1)
                                if type == 'Ручная(Сборка на ЦФЗ)':
                                    who.append(uuid2)
                                if type == 'Ручная(На этапе доставки)':
                                    who.append(uuid3)
                            result = {'Номер заказа': df[0].tolist(), 'Время заказа': df[1].tolist(), 'order_id': df[2].tolist(), 'Тип корректировки': type_update,
                                      'Кто скорректировал': who, 'product_id': df[10].tolist()}
                            res = pd.DataFrame(result)
                            connection1 = psycopg2.connect(user=login,
                                                          password=password,
                                                          host="patroni-06.samokat.io",
                                                          port="5434",
                                                          dbname="employee_profiles_backend")
                            cursor1 = connection1.cursor()
                            profile_id = tuple(list(filter(None, res['Кто скорректировал'].values)))
                            count = int(len(profile_id))
                            employees = []
                            if profile_id == ():
                                pass
                            if count == 1:
                                profile_id = profile_id[0]
                                cursor1.execute(f"SELECT profile_id, full_name FROM profile WHERE profile_id = '{profile_id}'")
                                employees.extend(cursor1.fetchall())
                            if count > 1:
                                cursor1.execute(f'SELECT profile_id, full_name FROM profile WHERE profile_id in {profile_id}')
                                employees.extend(cursor1.fetchall())
                            who = res['Кто скорректировал'].to_list()
                            who_update = []
                            for id in who:
                                for employee in employees:
                                    if id == employee[0]:
                                        who_update.append(employee[1])
                                if id == '':
                                    who_update.append(id)
                            products = [i for i in df[10].values]
                            search_json = {"productIds": [i for i in df[10].values if i is not None]}
                            response = requests.post('https://ds-metadata.samokat.ru/products/by-ids', json=search_json)
                            response_json = response.json()
                            products_name = []
                            for id in df[10].tolist():
                                for product in response_json:
                                    if id == product['productId']:
                                        products_name.append(product['administrativeName'])
                                if id is None:
                                    products_name.append('')
                            result = {'Номер заказа': df[0].tolist(), 'Время заказа': df[1].tolist(),
                                      'order_id': df[2].tolist(), 'Тип корректировки': type_update,
                                      'Кто скорректировал': who_update, 'product_id': df[10].tolist(), 'Продукт': products_name}
                            res = pd.DataFrame(result)
                            str_current_datetime = str(datetime.now()).replace(':', '-')
                            file_name = 'Отчет по корректировкам ' + str_current_datetime + '.xlsx'
                            writer = pd.ExcelWriter(file_name)
                            res.to_excel(writer, index=False)
                            writer.close()
                            wb = xw.Book(file_name)
                            sheet = wb.sheets[0]
                            sheet.range('A:A').column_width = 15
                            sheet.range('B:B').column_width = 15
                            sheet.range('C:C').column_width = 40
                            sheet.range('D:D').column_width = 25
                            sheet.range('E:E').column_width = 40
                            sheet.range('F:F').column_width = 40
                            sheet.range('G:G').column_width = 70
                            wb.save()
                            wb.close()
                            self.save_log('Готово, создан файл: ' + file_name)
            except (Exception, Error) as error:
                print("Ошибка при работе с PostgreSQL", error)


    def start_app10(self):
        self.logs.clear()
        vpn = self.vpn_on()
        if vpn is True:
            try:
                date = self.dateTimeEdit.text()
                date2 = self.dateTimeEdit_2.text()
                store_id = self.plainTextEdit_3.toPlainText()
                if store_id == ['']:
                    self.save_log('Вы не ввели store id')
                else:
                    login = Cache.load("login")
                    password = Cache.load("password")
                    if login is None:
                        self.save_log('Вы не авторизовались')
                    else:
                        connection = psycopg2.connect(user=login,
                                                      password=password,
                                                      host="patroni-17.samokat.io",
                                                      port="5434",
                                                      dbname="order_history")
                        cursor = connection.cursor()
                        date = date.split(' ')
                        if date[1] == '00:00:00':
                            date[1] = '00:00:01'
                        date = [date[0] + ' ' + date[1]]
                        date2 = [date2]
                        df = pd.DataFrame({
                            'datefrom': date,
                            'datetill': date2
                        }).astype({'datefrom': 'datetime64', 'datetill': 'datetime64'})
                        dfrom, dtill = df.at[0, 'datefrom'], df.at[0, 'datetill']
                        df1 = pd.DataFrame({'date': pd.date_range(dfrom, dtill, freq='D', normalize=True)}).assign(
                            datefrom=lambda x: x['date'])
                        df1['datetill'] = df1.datefrom + pd.Timedelta(1, unit='d') - pd.Timedelta(1, unit='s')
                        df1.at[df1.iloc[0].name, 'datefrom'], df1.at[df1.iloc[-1].name, 'datetill'] = dfrom, dtill
                        result = []
                        for date1, date2 in zip(df1['datefrom'].tolist(), df1['datetill'].tolist()):
                            query = f'WITH orders AS (SELECT order_id, order_line_changed  FROM order_history WHERE store_id = \'{store_id}\' and order_line_changed = true and created_date_time between \'{date1}\' and \'{date2})\'), date_created AS (SELECT change_date as created_date, order_id FROM order_status WHERE order_id in (SELECT order_id FROM orders) and order_status_id = 3), date_picking AS (SELECT change_date as picking_date, order_id FROM order_status WHERE order_id in (SELECT order_id FROM orders) and order_status_id = 4), date_picking_hub AS (SELECT change_date as picking_date_hub, order_id FROM order_status WHERE order_id in (SELECT order_id FROM orders) and order_status_id = 9), date_picked AS (SELECT change_date as picked_date,order_id FROM order_status WHERE order_id in (SELECT order_id FROM orders) and order_status_id = 5), date_picked_hub AS (SELECT change_date as picked_date_hub,order_id FROM order_status WHERE order_id in (SELECT order_id FROM orders) and order_status_id = 10), date_change AS (SELECT created_date_time as change_date, order_id FROM order_change WHERE order_id in (SELECT order_id FROM orders)), hub_picker AS (SELECT order_id, picker_id FROM distribution_center_picking_info WHERE order_id in (SELECT order_id FROM orders)), cfz_picker AS (SELECT order_id, picker_uuid FROM picking_info WHERE order_id in (SELECT order_id FROM orders)), cfz_deliveryman AS (SELECT order_id, deliveryman_uuid FROM delivery_info WHERE order_id in (SELECT order_id FROM orders)), order_number AS (SELECT order_id, display_number FROM order_history WHERE order_id in (SELECT order_id FROM orders)), product AS (SELECT oc.created_date_time as change_dates, olc.product_id FROM order_change oc JOIN order_line_change olc ON olc.order_change_id = oc.id WHERE oc.order_id in (SELECT order_id FROM orders)), accepted AS(SELECT order_id, product_id, quantity FROM order_line ol JOIN accepted_order_line aol ON aol.order_line_id = ol.id WHERE aol.order_id in (SELECT order_id FROM orders)), actual AS(SELECT order_id, product_id, quantity FROM order_line ol JOIN actual_order_line aol ON aol.order_line_id = ol.id WHERE aol.order_id in (SELECT order_id FROM orders)), changes AS(SELECT accepted.order_id, accepted.product_id, CASE WHEN accepted.quantity > actual.quantity THEN true ELSE false END AS change FROM accepted JOIN actual ON actual.product_id = accepted.product_id and actual.order_id = accepted.order_id), changes_2 AS (SELECT order_id, product_id FROM changes WHERE change is true), result AS(SELECT order_number.display_number AS "Номер заказа", to_char(date_created.created_date, \'yyyy-mm-dd hh24:mi\') AS "Время заказа", date_created.order_id, CASE WHEN change_date < picking_date_hub THEN true WHEN change_date < picking_date and picking_date_hub is NULL THEN true WHEN change_date is NULL and order_line_changed = true THEN true WHEN change_date is not NULL and order_line_changed = true and picking_date_hub is NULL and picking_date is NULL THEN true ELSE false END "Автокорректировка", CASE WHEN change_date > picking_date_hub and change_date < picked_date_hub THEN true ELSE false END "Сборка на Хабе", hub_picker.picker_id, CASE WHEN change_date > picking_date and change_date < picked_date and picking_date_hub is NULL THEN true ELSE false END "Сборка на ЦФЗ", cfz_picker.picker_uuid, CASE WHEN change_date > picked_date THEN true ELSE false END "Доставка", cfz_deliveryman.deliveryman_uuid, product.product_id AS "Продукт", changes_2.product_id AS "Продукт2" FROM date_created LEFT JOIN date_picking ON date_created.order_id = date_picking.order_id LEFT JOIN date_picking_hub ON date_picking_hub.order_id = date_created.order_id LEFT JOIN date_picked ON date_created.order_id = date_picked.order_id LEFT JOIN date_picked_hub ON date_picked_hub.order_id = date_created.order_id LEFT JOIN date_change ON date_change.order_id = date_created.order_id LEFT JOIN orders ON orders.order_id = date_created.order_id LEFT JOIN hub_picker ON hub_picker.order_id = date_created.order_id LEFT JOIN cfz_picker ON cfz_picker.order_id = date_created.order_id LEFT JOIN cfz_deliveryman ON cfz_deliveryman.order_id = date_created.order_id LEFT JOIN product ON product.change_dates = date_change.change_date LEFT JOIN order_number ON order_number.order_id = date_created.order_id LEFT JOIN changes_2 ON changes_2.order_id = date_created.order_id GROUP BY "Время заказа", date_created.order_id, "Автокорректировка", "Сборка на Хабе","Сборка на ЦФЗ", "Доставка", hub_picker.picker_id, cfz_picker.picker_uuid,cfz_deliveryman.deliveryman_uuid, "Продукт", "Номер заказа", "Продукт2" ORDER BY date_created.order_id ASC, "Сборка на Хабе" DESC) SELECT "Номер заказа", "Время заказа", order_id, "Автокорректировка", "Сборка на Хабе", picker_id, "Сборка на ЦФЗ", picker_uuid, "Доставка", deliveryman_uuid, CASE  WHEN "Автокорректировка" is true and "Продукт2" is NULL THEN "Продукт" WHEN "Автокорректировка" is true and "Продукт2" is not NULL THEN "Продукт2" WHEN "Автокорректировка" is false THEN "Продукт" END "Продукт" FROM result'
                            cursor.execute(query)
                            result.extend(cursor.fetchall())
                        if not result:
                            self.save_log('Корректировки отсутствуют')
                        else:
                            df = pd.DataFrame(result)
                            type_update = []
                            for a, b, c, d in zip(df[3].tolist(), df[4].tolist(), df[6].tolist(), df[8].tolist()):
                                if a is True:
                                    type_update.append('Автокорректировка')
                                if b is True:
                                    type_update.append('Ручная(Сборка на ХАБе)')
                                if c is True:
                                    type_update.append('Ручная(Сборка на ЦФЗ)')
                                if d is True:
                                    type_update.append('Ручная(На этапе доставки)')

                            result = {'Номер заказа': df[0].tolist(), 'Время заказа': df[1].tolist(), 'order_id': df[2].tolist(), 'Тип корректировки': type_update,
                                      'product_id': df[10].tolist()}
                            res = pd.DataFrame(result)
                            who = []
                            for type, uuid1, uuid2, uuid3 in zip(res['Тип корректировки'].tolist(), df[5].tolist(), df[7].tolist(),
                                                                 df[9].tolist()):
                                if type == 'Автокорректировка':
                                    who.append('')
                                if type == 'Ручная(Сборка на ХАБе)':
                                    who.append(uuid1)
                                if type == 'Ручная(Сборка на ЦФЗ)':
                                    who.append(uuid2)
                                if type == 'Ручная(На этапе доставки)':
                                    who.append(uuid3)
                            result = {'Номер заказа': df[0].tolist(), 'Время заказа': df[1].tolist(), 'order_id': df[2].tolist(), 'Тип корректировки': type_update,
                                      'Кто скорректировал': who, 'product_id': df[10].tolist()}
                            res = pd.DataFrame(result)
                            connection1 = psycopg2.connect(user=login,
                                                           password=password,
                                                           host="patroni-06.samokat.io",
                                                           port="5434",
                                                           dbname="employee_profiles_backend")
                            cursor1 = connection1.cursor()
                            profile_id = tuple(list(filter(None, res['Кто скорректировал'].values)))
                            count = int(len(profile_id))
                            employees = []
                            if profile_id == ():
                                pass
                            if count == 1:
                                profile_id = profile_id[0]
                                cursor1.execute(f"SELECT profile_id, full_name FROM profile WHERE profile_id = '{profile_id}'")
                                employees.extend(cursor1.fetchall())
                            if count > 1:
                                cursor1.execute(f'SELECT profile_id, full_name FROM profile WHERE profile_id in {profile_id}')
                                employees.extend(cursor1.fetchall())
                            who = res['Кто скорректировал'].to_list()
                            who_update = []
                            for id in who:
                                for employee in employees:
                                    if id == employee[0]:
                                        who_update.append(employee[1])
                                if id == '':
                                    who_update.append(id)
                            search_json = {"productIds": [i for i in df[10].values if i is not None]}
                            response = requests.post('https://ds-metadata.samokat.ru/products/by-ids', json=search_json)
                            response_json = response.json()
                            products_name = []
                            for id in df[10].tolist():
                                for product in response_json:
                                    if id == product['productId']:
                                        products_name.append(product['administrativeName'])
                                if id is None:
                                    products_name.append('')
                            result = {'Номер заказа': df[0].tolist(), 'Время заказа': df[1].tolist(), 'order_id': df[2].tolist(), 'Тип корректировки': type_update,
                                      'Кто скорректировал': who_update, 'product_id': df[10].tolist(), 'Продукт': products_name}
                            res = pd.DataFrame(result)
                            str_current_datetime = str(datetime.now()).replace(':', '-')
                            file_name = 'Отчет по корректировкам ' + str_current_datetime + '.xlsx'
                            writer = pd.ExcelWriter(file_name)
                            res.to_excel(writer, index=False)
                            writer.close()
                            wb = xw.Book(file_name)
                            sheet = wb.sheets[0]
                            sheet.range('A:A').column_width = 15
                            sheet.range('B:B').column_width = 15
                            sheet.range('C:C').column_width = 40
                            sheet.range('D:D').column_width = 25
                            sheet.range('E:E').column_width = 40
                            sheet.range('F:F').column_width = 40
                            sheet.range('G:G').column_width = 70
                            wb.save()
                            wb.close()
                            self.save_log('Готово, создан файл: ' + file_name)
            except (Exception, Error) as error:
                print("Ошибка при работе с PostgreSQL", error)


    def start_app11(self):
        self.logs.clear()
        vpn = self.vpn_on()
        if vpn is True:
            try:
                find_feature =  self.plainTextEdit_4.toPlainText()
                guids = self.plainTextEdit_5.toPlainText().split('\n')
                invalid_guids, features_enabled, features_disabled = [], [], []
                guids = [guid for guid in guids if len(guid) == 36]
                if len(guids) == 0:
                    self.save_log('Вы не ввели guid ЦФЗ')
                else:
                    self.save_log('Вы ввели ' + str(len(guids)) + ' guid ЦФЗ')
                    for guid in guids:
                        if len(guid) == 36:
                            cfz = requests.get(f'https://ds-warehouse.samokat.ru/warehouses/{guid}/settings')
                            cfz = cfz.json()
                            features = cfz['value']['features']
                            if find_feature in features:
                                features_enabled.append(guid)
                            else:
                                features_disabled.append(guid)
                        else:
                            invalid_guids.append(guid)

                    count_enabled = len(features_enabled)
                    count_disabled = len(features_disabled)
                    str_current_datetime = str(datetime.now()).replace(':', '-')
                    file_name = 'cfz_feature ' + str_current_datetime + '.json'
                    with open(file_name, 'w', encoding="utf-8") as f:
                        f.write(f'Фича - {find_feature}\n')
                        f.write(f'Включена на : {count_enabled} ЦФЗ\n')
                        for fe in features_enabled:
                            f.write(f'{fe}\n')
                        f.write(f'Выключена на : {count_disabled} ЦФЗ\n')
                        for fd in features_disabled:
                            f.write(f'{fd}\n')
                        if len(invalid_guids) > 0:
                            f.write(f'Некорректные guid:\n')
                            for invalid in invalid_guids:
                                f.write(invalid)
                    self.save_log('Готово, создан файл: ' + file_name)
            except Exception:
                self.save_log('Вы не авторизовались')


    def get_employee(self, profileId):
        url = f'https://employee-profiles-backend.samokat.ru/profiles/{profileId}'
        response = requests.get(url)
        response = response.json()
        profile = response["name"]["firstName"] + ' ' + response["name"]["lastName"] + ' ' + response["name"][
            "middleName"]
        return profile

    def start_app14(self):
        self.logs.clear()
        vpn = self.vpn_on()
        if vpn is True:
            try:
                packages = self.plainTextEdit_6.toPlainText().split('\n')
                if len(packages) == 0:
                    self.save_log('Вы не ввели номера посылок')
                else:
                    self.save_log('Вы ввели ' + str(len(packages)) + ' RP')
                    NOT_RECEIVED = []
                    RECEIVED = {}
                    for package in packages:
                        package_search = {"externalIds": [package],"pageSize": 10, "pageNumber": 0}
                        url_packages = 'https://shipment-package.samokat.ru/v1/internal/shipments/short-filter'
                        url_tasks = 'https://dms-supply.samokat.ru/supply/support/tasks'
                        response_package = requests.post(url_packages, json=package_search)
                        response_package = response_package.json()
                        store = response_package['value']['content'][0]['storeId']
                        barcode = response_package['value']['content'][0]['packages'][0]['barcode']
                        logs = response_package['value']['content'][0]['packages'][0]['log']
                        logs_arr,dates_delivery, dates_refunds = [], [], []
                        for log in logs:
                            logs_arr.append(log['status'])
                        if "NOT_RECEIVED" in logs_arr and "READY_FOR_DELIVERY" not in logs_arr:
                            NOT_RECEIVED.append(package)
                        else:
                            for log in logs:
                                if log['status'] == "READY_FOR_DELIVERY":
                                    dates_delivery.append(log['timeAt'])
                            for date in dates_delivery:
                                date_from = str(date).split('T')[0] + "T00:00:00Z"
                                date_to = str(date).split('T')[0] + "T23:59:59Z"
                                tasks_search = {"storeId": store, "from": date_from,
                                                "to": date_to,
                                                "acceptanceStatuses": ["COMPLETED"], "documentTypes": ["PACKAGES"], "offset": 0,
                                                "limit": 10}
                                response_tasks = requests.get(url_tasks, params=tasks_search)
                                response_tasks = response_tasks.json()
                                for task in response_tasks['value']:
                                    task_search = task['mobileApplicationView']['taskId']
                                    url_task = f'https://dms-supply.samokat.ru/supply/support/tasks/{task_search}'
                                    response_task = requests.get(url_task)
                                    response_task = response_task.json()
                                    for package_accepted in response_task['value']['mobileApplicationView']['packages']:
                                        if package_accepted['barcode'] == barcode and package_accepted['isAccepted'] is True:
                                            for event in response_task['value']['eventLog']:
                                                if event['event'] == "completed":
                                                    profile = event["userId"]
                                                    user = self.get_employee(profile)
                                                    res = {"RECEIVED": f'Принял(а) {user} в {event["timestamp"]}'}
                                                    RECEIVED[package] = res
                            for log in logs:
                                if log['status'] == "READY_FOR_REFUND":
                                    for log in logs:
                                        if log['status'] == "READY_FOR_REFUND":
                                            dates_refunds.append(log['timeAt'])
                                    for date in dates_refunds:
                                        date_from = str(date).split('T')[0] + "T00:00:00Z"
                                        date_to = str(date).split('T')[0] + "T23:59:59Z"
                                        tasks_search = {"storeId": store, "from": date_from,
                                                        "to": date_to,
                                                        "acceptanceStatuses": ["COMPLETED"], "documentTypes": ["PACKAGES_REFUND"],
                                                        "offset": 0,
                                                        "limit": 10}
                                        response_tasks = requests.get(url_tasks, params=tasks_search)
                                        response_tasks = response_tasks.json()
                                    for task in response_tasks['value']:
                                        task_search = task['mobileApplicationView']['taskId']
                                        url_task = f'https://dms-supply.samokat.ru/supply/support/tasks/{task_search}'
                                        response_task = requests.get(url_task)
                                        response_task = response_task.json()
                                        for package_accepted in response_task['value']['mobileApplicationView']['packages']:
                                            if package_accepted['barcode'] == barcode and package_accepted['isAccepted'] == True:
                                                for event in response_task['value']['eventLog']:
                                                    if event['event'] == "completed":
                                                        user = self.get_employee(event["userId"])
                                                        res = f'Вернул(а) {user} в {event["timestamp"]}'
                                                        RECEIVED[package]["REFUND"] = res

                str_current_datetime = str(datetime.now()).replace(':', '-')
                file_name = 'search_package ' + str_current_datetime + '.json'
                with open(file_name, 'w', encoding="utf-8") as f:
                    f.write(f'Не принимали: {NOT_RECEIVED}\n')
                    f.write(json.dumps(RECEIVED, indent=4, ensure_ascii=False))
                    self.save_log('Готово, создан файл: ' + file_name)

            except Exception:
                self.save_log('Возникла ошибка')

    def start_app15(self):
        body = self.plainTextEdit_7.toPlainText()
        print(body)
        if body == '':
            self.save_log('Необходимо вставить лог')
        else:
            try:
                body = body.split('payload=[Payload(')
                body = body[1].split('=')
                print(body)
                dateRecruitment = body[1].split(',')[0]
                print(dateRecruitment)
                dateDismiss = body[2].split(',')[0]
                print(dateDismiss)
                jobTitle = body[4].split(')')[0]
                print(jobTitle)
                guid = body[6].split(',')[0]
                print(guid)
                email = body[7].split(',')[0]
                print(email)
                lastName = body[8].split(',')[0]
                print(lastName)
                inn = body[9].split(',')[0]
                print(inn)
                firstName = body[10].split(',')[0]
                print(firstName)
                phoneNumber = "".join(c for c in body[11].split(',')[0] if c.isdecimal())
                print(phoneNumber)
                patronymic = "".join(c for c in body[12].split(',')[0] if c.isalnum())
                print(patronymic)
                city = body[14].split(')')[0]
                print(city)
                subUnit = body[18].split(',')[0].replace('))',')')
                print(subUnit)
                leader_email = body[20].split(')')[0]
                print(leader_email)
                body_postman = {
                                "guid": guid,
                                "inn": inn,
                                "addressEP": email,
                                "imya": firstName,
                                "familia": lastName,
                                "otchestvo": patronymic,
                                "rukovoditelAdresEP": leader_email,
                                "cityCaption": city,
                                "podrazdelenie": subUnit,
                                "post": jobTitle,
                                "dataPriema": dateRecruitment,
                                "dataUvolneniya": dateDismiss,
                                "phone": phoneNumber
                                }
                print(body_postman)
                str_current_datetime = str(datetime.now()).replace(':', '-')
                file_name = 'restart(staff) ' + str_current_datetime + '.json'
                with open(file_name, 'w', encoding="utf-8") as f:
                    f.write(json.dumps(body_postman, indent=4, ensure_ascii=False))
                    self.save_log('Готово, создан файл: ' + file_name)
            except Exception:
                self.save_log("Возникла ошибка")



"""
    def start_app12(self):
        self.logs.clear()
        vpn = self.vpn_on()
        if vpn is True:
            try:
                token = Cache.load("token")
                guids = self.plainTextEdit.toPlainText().split('\n')
                guids = [guid for guid in guids if len(guid) == 36]
                if len(guids) == 0:
                    self.save_log('Вы не ввели guid заказа')
                else:
                    self.save_log('Вы ввели ' + str(len(guids)) + ' guid заказов')
                    str_current_datetime = str(datetime.now()).replace(':', '-')
                    file_name = 'cfz_settings ' + str_current_datetime + '.json'
                    with open(file_name, 'w', encoding="utf-8") as f:
                        for guid in guids:
                            if len(guid) == 36:
                                receipts_search = {"orderId": guid}
                                header = {'Authorization': 'Bearer ' + token}
                                url_receipts = 'https://smk-supportpaymentgw.samokat.ru/receipt/cash-registers/find'
                                receipt = requests.get(url_receipts, headers=header, params=receipts_search)
                                receipt = receipt.json()
                                receipt_url =
                                f.write(f'Заказ: {guid}, ссылка на чек: {receipt_url}\n')
                            else:
                                continue
                    self.save_log('Готово, создан файл: ' + file_name)
            except Exception:
                self.save_log('Вы не авторизовались')
"""
